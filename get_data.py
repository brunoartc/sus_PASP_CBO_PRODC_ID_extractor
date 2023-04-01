import requests
import json

from typing import Any


import openpyxl

wb_obj = openpyxl.Workbook()

PROGRAMA_ID = 6

def extract_from_municipio(municipio_id:int = 355030, uf_ibge= 35) -> Any:

    result = []

    dat_acad = requests.get(f'https://sismobcidadao.saude.gov.br/api/public/obras?size=10&ufIbge={uf_ibge}&municipioIbge={municipio_id}&programaId={PROGRAMA_ID}&sgUf=&noMunicipio=', verify=False)
    data_acad = json.loads(dat_acad.content)
    for acad in data_acad['content']:
        result.append(acad)
    return result

def extract_obra_id(obra_id:int = 18401) -> Any:

    dat_acad = requests.get(f'https://sismobcidadao.saude.gov.br/api/public/obras/{obra_id}', verify=False)
    data_acad = json.loads(dat_acad.content)
    return data_acad


def data_json_to_tsv_get_obra_info(data_program) -> str:
    out_csv = []
    out_csv_li = []

    if len(data_program) == 0:
        return "", []

    #header = ""

    #for key in data_program[0].keys():
    #    header += f"{key}\t"


    #for key in extract_obra_id(data_program[0]['propostaId']).keys():
    #    header += f"{key}\t"
    
    #out_csv.append(header[:-1])

    for _data in data_program:
        line_data = ""
        line_data_li = []
        for acad_value in _data.values():
            line_data += f"{acad_value}\t"
            line_data_li.append(acad_value)

        data_obra = extract_obra_id(_data['propostaId'])

        for kk,data_acad_value in data_obra.items():
            if kk == "gruposFotografias":
                continue
            if type(data_acad_value) == type(""):
                data = data_acad_value.replace("\n", "").replace(",", " ")
            else:
                data = data_acad_value
            line_data_li.append(data)
            line_data += f"{data}\t"

        out_csv.append(line_data[:-1])
        out_csv_li.append(line_data_li)
    return "\n".join(out_csv), out_csv_li 
        

def data_json_to_csv(data_program) -> str:
    out_csv = []

    if len(data_program) == 0:
        return ""

    header = ""

    for key in data_program[0].keys():
        header += f"{key},"
    
    out_csv.append(header[:-1])

    for _data in data_program:
        line_data = ""
        for acad_value in _data.values():
            line_data += f"{acad_value},"
        out_csv.append(line_data[:-1])
    return "\n".join(out_csv)

ufs = [11,    12,    13,    14,    15,    16,   17,    21,    22,    23,    24,    25,    26,    27,    28,    29,    31,    32,    33,    35,    41,    42,    43,    50,    51,    52,    53]

ccc = 0



countt = 0

csv_out = "propostaId	numeroProposta	situacaoObra	coSituacaoObra	uf	municipio	bairro	novoBairro	tipoObra	coTipoObra	tipoRecurso	programa	redePrograma	cnes	nomeEstabelecimento	nuLatitude	nuLongitude	vlPercentualExecutado	vlProposta	coSeqProposta	sgUf	noMunicipio	noMunicipioAcentuado	nuCnpj	noPadronizadoEntidade	coEsferaAdministrativa	dsEsferaAdministrativa	nuProposta	coPrograma	dsPrograma	dsRedePrograma	coTipoObra	dsTipoObra	coTipoRecurso	dsTipoRecurso	dsTipoRecursoFiltro	dsPortePrograma	dtCadastro	nuPortaria	dtPortaria	vlProposta	nuAnoReferencia	dsEtapaProposta	dsSituacaoObra	dtMudancaSituacao	dsJustificativa	nuCep	dsLogradouro	nuEndereco	dsComplemento	dsBairro	coBairro	noBairro	nuLatitude	nuLongitude	coUnidade	coCnes	noEstabelecimentoCnes	noEstabelecimentoProposta	coFaseProjeto	coSituacaoObra	dsFaseProjeto	dtInicioProjeto	dtPrevistaInicioProjeto	dtPrevistaConclusaoProjeto	dtConclusaoProjeto	coTipoExecucaoProjeto	dsTipoFormaExecucaoProjeto	vlPercentualExecutado	dtInicioObra	dtProvavelExecucao	dtProvavelConclusaoFinal	dtExecucao	dtConclusaoFinal	stAditivoContratual	stPossuiEtapaFuncionamento	vlTotalContrato	dtPrimeiraParcela	vlPrimeraParcela	dtSegundaParcela	vlSegundaParcela	dtTerceiraParcela	vlTerceiraParcela	dtQuartaParcela	vlQuartaParcela	dtAtualizacao	dtOrdemServico	dtPrevistaInicioFuncionamento	dtInicioFuncionamento	dtPrevistaInauguracao	dtInauguracao	nuCnes	gruposFotografias	empresas\n"


sheet_obj = wb_obj.active

for header in csv_out.split("\t"):
    sheet_obj.cell(row = 1, column = ccc + 1).value = header
    ccc += 1



for uf_id in ufs:
    UF_IBGE = 35


    dat_mun = requests.get(f'https://sismobcidadao.saude.gov.br/api/public/endereco/municipios?pageSize=10&ufIbge={uf_id}', verify=False)
    data_municipios = json.loads(dat_mun.content)

    ids_municipios = [int(i['ibge']) for i in data_municipios]





    

    for municipio_id in ids_municipios:

        countt += 1

        print(f".{countt}/{len(ids_municipios)} ------------- {ufs.index(uf_id)}/{len(ufs)}")

        date_municipio = extract_from_municipio(municipio_id, uf_id)

        csv_out_o, csv_out_li = data_json_to_tsv_get_obra_info(date_municipio)
        csv_out += csv_out_o
        for __data in csv_out_li:
            for i in range(len(__data)):
                if type(__data[i]) != list and type(__data[i]) != dict:
                    sheet_obj.cell(row = countt+1, column = i+1).value = __data[i]

with open("out.tsv", "w") as text_file:
    text_file.write(csv_out)

wb_obj.save("sample.xlsx")